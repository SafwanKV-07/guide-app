from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import SQLAlchemyError
from models import db, Category, Subcategory, Update
from search import search_index, Trie
import logging
from datetime import datetime, timedelta
import traceback

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


def safe_str(value):
    """Safely convert any value to a string"""
    if isinstance(value, tuple):
        return " ".join(str(v) for v in value)
    return str(value) if value is not None else ""


def get_value_from_merge(sheet, row, col):
    """Get value from a potentially merged cell and ensure it's a string"""
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            return safe_str(value)
    value = sheet.cell(row, col).value
    return safe_str(value)


def load_data(app):
    excel_file_path = app.config["EXCEL_FILE_PATH"]

    try:
        # Load the workbook
        wb = load_workbook(filename=excel_file_path, read_only=False)

        with app.app_context():
            # Clear existing data
            db.session.query(Update).delete()
            db.session.query(Subcategory).delete()
            db.session.query(Category).delete()
            db.session.commit()

            # Reset the search index
            search_index.documents.clear()
            search_index.inverted_index.clear()
            search_index.trie = Trie()

            # Process main data sheets
            for sheet_name in wb.sheetnames:
                if (
                    sheet_name != "Updates"
                ):  # Assuming "Updates" is the name of your updates sheet
                    sheet = wb[sheet_name]
                    process_main_sheet(sheet)

            # Process Updates sheet
            if "Updates" in wb.sheetnames:
                updates_sheet = wb["Updates"]
                process_updates_sheet(updates_sheet)

            db.session.commit()
            logger.info("Data loaded successfully and search index updated.")
            return "Data loaded successfully and search index updated."

    except FileNotFoundError:
        logger.error(f"Excel file not found at {excel_file_path}")
        return f"Error: Excel file not found at {excel_file_path}"
    except SQLAlchemyError as e:
        db.session.rollback()
        logger.error(f"Database error occurred: {str(e)}")
        return f"Error: Database error occurred: {str(e)}"
    except Exception as e:
        logger.error(f"An unexpected error occurred: {str(e)}")
        logger.error(traceback.format_exc())
        return f"Error: An unexpected error occurred: {str(e)}"
    finally:
        if "wb" in locals():
            wb.close()


def process_main_sheet(sheet):
    headers = [safe_str(cell.value) for cell in sheet[1]]
    logger.debug(f"Headers for sheet {sheet.title}: {headers}")

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            row_data = {
                header: get_value_from_merge(sheet, row_index, col_index + 1)
                for col_index, header in enumerate(headers)
            }

            main_folder = row_data.get("Main Folder", "").strip()
            sub_folder = row_data.get("Sub Folder", "").strip()

            if not main_folder and not sub_folder:
                continue

            category = Category.query.filter_by(name=main_folder).first()
            if not category:
                category = Category(name=main_folder)
                db.session.add(category)

            if sub_folder:
                subcategory = Subcategory(
                    name=sub_folder,
                    document_type=row_data.get("Document Type", ""),
                    identification_rules=row_data.get(
                        "Document Type Identification Rules", ""
                    ),
                    supporting_information=row_data.get("Supporting Information", ""),
                    category=category,
                )
                db.session.add(subcategory)
                search_index.add_document(subcategory)

        except Exception as e:
            logger.error(
                f"Error processing row {row_index} in sheet {sheet.title}: {str(e)}"
            )
            logger.error(traceback.format_exc())


def parse_date(date_string):
    if not date_string:
        return None
    try:
        return datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S").date()
    except ValueError:
        try:
            return datetime.strptime(date_string, "%d-%m-%Y").date()
        except ValueError:
            logger.warning(f"Unable to parse date: {date_string}")
            return None


def process_updates_sheet(sheet):
    headers = [safe_str(cell.value) for cell in sheet[1]]
    logger.debug(f"Update Headers: {headers}")

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            row_data = {
                header: get_value_from_merge(sheet, row_index, col_index + 1)
                for col_index, header in enumerate(headers)
            }

            date_value = parse_date(row_data.get("Date", ""))
            if date_value is None:
                logger.warning(f"Skipping row {row_index} due to invalid date")
                continue

            update = Update(
                reference=row_data.get("Reference", ""),
                date=date_value,
                main_folder=row_data.get("Main Folder", ""),
                category=row_data.get("Category", ""),
                description=row_data.get("Description", ""),
            )
            db.session.add(update)

        except Exception as e:
            logger.error(f"Error processing row {row_index} in Updates sheet: {str(e)}")
            logger.error(traceback.format_exc())


def reload_data(app):
    logger.info("Reloading data...")
    result = load_data(app)
    logger.info("Data reload complete.")
    return result
